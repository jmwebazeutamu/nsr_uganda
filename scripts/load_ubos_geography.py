"""Loader for the UBOS administrative hierarchy from the supplied workbook.

Sprint 0 item 3 per CLAUDE.md and SAD §11.4.

Input file shape (header row, then 10,854 data rows in the May 2026 supply):
    District_code | District_Name | County_code | County_Name |
    Subcounty_code | Subcounty_Name | Parish_code | Parish_Name

Coverage gap: this source only carries four levels (district to parish).
The schema supports seven (region, sub_region, district, county, sub_county,
parish, village). Region and sub_region come from a separate UBOS sheet
that has not been supplied yet; village polygons depend on DQA-O-03 (UBOS
village polygons availability). The loader leaves those levels empty.

Codes are composed hierarchically so they are globally unique per level
(the raw codes in the file are scoped within their parent, so subcounty
code "01" appears thousands of times):

    District:    "101"
    County:      "101.1"
    Subcounty:   "101.1.01"
    Parish:      "101.1.01.01"

Idempotent: the unique constraint on (level, code, effective_from) lets
re-runs skip rows already present.

Usage:
    .venv/bin/python scripts/load_ubos_geography.py \\
        /Users/johnsonmwebaze/Downloads/Goegraphy_final_with_codes.xlsx

Options:
    --effective-from YYYY-MM-DD   Default 2026-01-01.
    --dry-run                     Print counts; no DB writes.
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import date
from pathlib import Path

import django

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "nsr_mis.settings")
django.setup()

from apps.reference_data.models import GeographicUnit  # noqa: E402
from django.db import transaction  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

HEADER_V1 = (
    "District_code", "District_Name", "County_code", "County_Name",
    "Subcounty_code", "Subcounty_Name", "Parish_code", "Parish_Name",
)

HEADER_V2 = (
    "District_code", "District_Name", "Region", "Subregion",
    "County_code", "County_Name", "Subcounty_code", "Subcounty_Name",
    "Parish_code", "Parish_Name",
)


def _slug(value: str) -> str:
    """Code suffix for the region/subregion levels which carry no UBOS code."""
    return value.strip().upper().replace(" ", "-").replace("/", "-")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("xlsx", help="Path to the UBOS hierarchy workbook")
    p.add_argument("--effective-from", default="2026-01-01",
                   help="Effective-from date for all rows (ISO). Default 2026-01-01.")
    p.add_argument("--dry-run", action="store_true", help="Print counts only; do not write.")
    return p.parse_args()


def _normalise(value) -> str:
    return str(value).strip() if value is not None else ""


def read_rows(path: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = tuple(next(it))
    if header == HEADER_V2:
        version = "v2"
    elif header == HEADER_V1:
        version = "v1"
    else:
        raise SystemExit(
            f"unexpected header: {header}\nexpected v1: {HEADER_V1}\nor v2: {HEADER_V2}"
        )
    rows: list[tuple[str, ...]] = []
    for row in it:
        if row is None or all(v is None for v in row):
            continue
        rows.append(tuple(_normalise(v) for v in row))
    return version, rows


def build_unique(version: str, rows: list[tuple[str, ...]]):
    """Walk the parish-level rows and collect unique nodes per level.

    Returns five dicts (regions/subregions empty for v1 input):
      regions:      code -> name
      subregions:   code -> (name, parent_region_code)
      districts:    code -> (name, parent_subregion_code | None)
      counties:     code -> (name, parent_district_code)
      subcounties:  code -> (name, parent_county_code)
      parishes:     code -> (name, parent_subcounty_code)
    """
    regions: dict[str, str] = {}
    subregions: dict[str, tuple[str, str]] = {}
    districts: dict[str, tuple[str, str | None]] = {}
    counties: dict[str, tuple[str, str]] = {}
    subcounties: dict[str, tuple[str, str]] = {}
    parishes: dict[str, tuple[str, str]] = {}

    for row in rows:
        if version == "v2":
            dc, dn, region_name, subregion_name, cc, cn, sc, sn, pc, pn = row
            r_code = f"R-{_slug(region_name)}" if region_name else None
            sr_code = (
                f"SR-{_slug(subregion_name)}-{_slug(region_name)}"
                if subregion_name and region_name else None
            )
        else:
            dc, dn, cc, cn, sc, sn, pc, pn = row
            region_name = subregion_name = ""
            r_code = sr_code = None

        if not (dc and cc and sc and pc):
            continue
        d_code = dc
        c_code = f"{dc}.{cc}"
        s_code = f"{dc}.{cc}.{sc}"
        p_code = f"{dc}.{cc}.{sc}.{pc}"

        if r_code:
            regions.setdefault(r_code, region_name)
        if sr_code:
            subregions.setdefault(sr_code, (subregion_name, r_code))
        districts.setdefault(d_code, (dn.title() if dn.isupper() else dn, sr_code))
        counties.setdefault(c_code, (cn.title() if cn.isupper() else cn, d_code))
        subcounties.setdefault(s_code, (sn.title() if sn.isupper() else sn, c_code))
        parishes.setdefault(p_code, (pn.title() if pn.isupper() else pn, s_code))

    return regions, subregions, districts, counties, subcounties, parishes


def load_level(level: str, mapping: dict, parent_lookup: dict[str, GeographicUnit] | None,
               effective_from: date) -> tuple[dict[str, GeographicUnit], int, int]:
    """Bulk-load one level. Idempotent: re-runs skip rows already present, but
    *back-fix* existing rows whose parent is NULL once a parent becomes
    available (e.g. v2 reload supplying a sub_region parent for districts
    previously loaded by v1)."""
    existing = {
        u.code: u for u in GeographicUnit.objects.filter(
            level=level, effective_from=effective_from,
        )
    }
    creates: list[GeographicUnit] = []
    fixed_parents = 0
    for code, payload in mapping.items():
        if isinstance(payload, tuple):
            name, parent_code = payload
            parent = parent_lookup.get(parent_code) if (parent_lookup and parent_code) else None
        else:
            name, parent = payload, None
        if code in existing:
            row = existing[code]
            if parent and row.parent_id is None:
                row.parent = parent
                row.save(update_fields=["parent"])
                fixed_parents += 1
            continue
        creates.append(GeographicUnit(
            level=level, code=code, name=name, parent=parent, effective_from=effective_from,
        ))

    if creates:
        GeographicUnit.objects.bulk_create(creates, batch_size=2000)

    refreshed = {
        u.code: u for u in GeographicUnit.objects.filter(
            level=level, effective_from=effective_from,
        )
    }
    return refreshed, len(creates), fixed_parents


def main() -> int:
    args = parse_args()
    eff = date.fromisoformat(args.effective_from)
    version, rows = read_rows(args.xlsx)
    print(f"read {len(rows)} parish rows ({version}) from {args.xlsx}")

    regions, subregions, districts, counties, subcounties, parishes = build_unique(version, rows)
    print(f"  unique regions:      {len(regions):>6}")
    print(f"  unique sub-regions:  {len(subregions):>6}")
    print(f"  unique districts:    {len(districts):>6}")
    print(f"  unique counties:     {len(counties):>6}")
    print(f"  unique sub-counties: {len(subcounties):>6}")
    print(f"  unique parishes:     {len(parishes):>6}")

    if args.dry_run:
        print("dry-run: nothing written")
        return 0

    with transaction.atomic():
        r_map,  r_new,  r_fix  = load_level("region",     regions,    None,   eff)
        sr_map, sr_new, sr_fix = load_level("sub_region", subregions, r_map,  eff)
        d_map,  d_new,  d_fix  = load_level("district",   districts,  sr_map, eff)
        c_map,  c_new,  c_fix  = load_level("county",     counties,   d_map,  eff)
        s_map,  s_new,  s_fix  = load_level("sub_county", subcounties, c_map, eff)
        p_map,  p_new,  p_fix  = load_level("parish",     parishes,   s_map,  eff)

    def _line(name, mp, new, fix):
        print(f"  {name:14s} in DB: {len(mp):>6}  (created {new:>5}, parent-backfilled {fix:>5})")

    _line("regions",     r_map,  r_new,  r_fix)
    _line("sub-regions", sr_map, sr_new, sr_fix)
    _line("districts",   d_map,  d_new,  d_fix)
    _line("counties",    c_map,  c_new,  c_fix)
    _line("sub-counties",s_map,  s_new,  s_fix)
    _line("parishes",    p_map,  p_new,  p_fix)

    print(f"\nGeographicUnit total rows: {GeographicUnit.objects.count()}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
