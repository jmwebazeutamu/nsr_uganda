"""US-118 — XLSForm export from FormVersion.

Walks the authored questionnaire (FormVersion → FormSection →
FormQuestion + ChoiceList catalogue) and emits a Kobo-compatible
XLSForm as raw .xlsx bytes. Output structure mirrors what
`k-forms/build_nsr_xlsform.py` produces — the legacy script is
the gold standard of Kobo-validity for this project.

Three sheets per the XLSForm spec (https://xlsform.org):

- `survey`   — top-level metadata rows (start/end/today/deviceid/
              username) auto-prepended, followed by one row per
              section opener + each question + section closer.
              Sections with `repeat_count` set are emitted as
              begin_repeat/end_repeat blocks (the household roster
              is the canonical case).
- `choices`  — one row per (ChoiceList, ChoiceOption) referenced
              by any select_one / select_multiple question in the
              survey sheet. Unreferenced lists are omitted to keep
              the xlsx small.
- `settings` — form_title / form_id / version / default_language /
              style. Matches the legacy Kobo-valid file:
              `English`, `theme-grid`, slugged form_id, YYYYMMDD
              version.

Pure-Python (openpyxl). No file system writes; caller is
responsible for streaming the returned bytes wherever they need
to go (admin download, MinIO upload, etc.).
"""

from __future__ import annotations

import io
import re

from openpyxl import Workbook

from .geo import (
    GEO_LEVELS_ROOT_TO_LEAF,
    GEO_PARENT_COLUMN,
    GEO_QUESTIONS,
    choice_filter_for,
    geo_options_for,
)
from .models import FormQuestion, FormSection, FormVersion

# XLSForm column orders — locked to the order the spec uses so a
# round-trip with off-the-shelf XLSForm tools doesn't shuffle
# columns.
SURVEY_COLS = (
    "type", "name", "label", "hint", "required", "relevant",
    "constraint", "constraint_message", "appearance",
    "choice_filter", "calculation", "repeat_count", "parameters",
)
# Choices sheet gets the standard 3 columns + an ancestor column
# per non-leaf geo level so cascading choice_filter expressions
# resolve (e.g. sub_region rows carry their region code in the
# `region` column). Non-geo lists leave the ancestor columns blank.
CHOICES_COLS = (
    "list_name", "name", "label",
    # Order is root→leaf-1: each child level's parent column.
    "region", "sub_region", "district", "county", "sub_county",
)
SETTINGS_COLS = (
    "form_title", "form_id", "version", "default_language", "style",
)

# Auto-prepended at the top of every export. Kobo's preview tolerates
# their absence but partner MIS tools sometimes don't — emitting them
# unconditionally keeps the output portable.
_METADATA_ROWS = (
    {"type": "start", "name": "start"},
    {"type": "end", "name": "end"},
    {"type": "today", "name": "today"},
    {"type": "deviceid", "name": "deviceid"},
    {"type": "username", "name": "username"},
)


def _slug(value: str) -> str:
    """Lowercase + underscore-only slug, suitable for form_id. Kobo
    accepts arbitrary strings but rejects ones with whitespace and
    odd punctuation — keep it conservative."""
    value = re.sub(r"[^A-Za-z0-9]+", "_", value.strip().lower())
    return re.sub(r"_+", "_", value).strip("_") or "form"


def _padded_choice_row(
    list_name: str, code: str, label: str,
    *, geo_parent_level: str = "", geo_parent_code: str = "",
) -> list[str]:
    """Build one row for the choices sheet, padded out to the full
    CHOICES_COLS width. If `geo_parent_level` is set, the parent's
    code is written into that ancestor column; all other ancestor
    columns are left blank. Non-geo lists pass both kwargs as ""
    and get fully-empty ancestor columns."""
    row = {"list_name": list_name, "name": code, "label": label}
    if geo_parent_level and geo_parent_code:
        row[geo_parent_level] = geo_parent_code
    return [row.get(c, "") for c in CHOICES_COLS]


def _flatten_parameters(value) -> str:
    """Project the FormQuestion.parameters dict to an XLSForm
    `parameters` cell string. Legacy script packs single values
    under the `_` key; richer dicts get key=value;... encoding."""
    if not value or not isinstance(value, dict):
        return ""
    if "_" in value and len(value) == 1:
        return str(value["_"])
    return ";".join(f"{k}={v}" for k, v in sorted(value.items()))


def _resolve_select_type(q: FormQuestion) -> str:
    """Pack the choice-list name back into the type cell for select
    questions. Three cases:

    1. The 6 legacy geo selects (region / subregion / …) use a list
       name derived from their GeographicUnit level — `select_one
       region`, `select_one sub_region`, etc. — and the choices
       sheet carries the matching rows + ancestor columns so
       cascading choice_filter expressions resolve at form runtime.
    2. Regular selects with a linked ChoiceList pack the list name
       from choice_list_ref.list_name.
    3. Broken selects (no choice_list_ref AND not a known geo
       question) fall back to `text` — emitting a bare `select_one`
       row fails Kobo's "Survey information not complete" validation
       so we'd rather degrade gracefully.
    """
    if q.type in ("select_one", "select_multiple"):
        if q.name in GEO_QUESTIONS:
            level = GEO_QUESTIONS[q.name][0]
            return f"{q.type} {level}"
        if q.choice_list_ref:
            return f"{q.type} {q.choice_list_ref.list_name}"
        return "text"  # fallback — see _question_row hint annotation
    return q.type


def _question_row(q: FormQuestion) -> dict:
    """Project a FormQuestion to its XLSForm survey row."""
    qtype = _resolve_select_type(q)
    hint = q.hint
    is_geo = q.name in GEO_QUESTIONS
    # Only flag the "missing choice list" hint for selects that
    # AREN'T being satisfied by the geo wiring.
    if q.type in ("select_one", "select_multiple") and not q.choice_list_ref and not is_geo:
        hint = (hint + " [missing choice list — exported as text]").strip()
    # Geo questions inherit appearance="minimal" if the author left
    # the field blank — matches the legacy Kobo form's UX where the
    # cascade is rendered as compact pickers, not radio lists.
    appearance = q.appearance
    if is_geo and not appearance:
        appearance = "minimal"
    return {
        "type": qtype,
        "name": q.name,
        "label": q.label,
        "hint": hint,
        "required": "yes" if q.required else "",
        "relevant": q.relevant_expression,
        "constraint": q.constraint_expression,
        "constraint_message": q.constraint_message,
        "appearance": appearance,
        "choice_filter": choice_filter_for(q.name) if is_geo else "",
        # US-S21-006 — XLSForm calculation cell. Required for
        # type=calculate; allowed (and sometimes used) on any other
        # type for client-side derived values. Kobo rejects calculate
        # rows whose cell is empty.
        "calculation": q.calculation or "",
        "repeat_count": q.repeat_count,
        "parameters": _flatten_parameters(q.parameters),
    }


def _section_open_row(s: FormSection) -> dict:
    """begin_group or begin_repeat for the section opener. Sections
    with a non-empty `repeat_count` (the household roster) emit as
    begin_repeat with the repeat_count column populated."""
    opener = "begin_repeat" if s.repeat_count else "begin_group"
    return {
        "type": opener,
        "name": s.name,
        "label": s.label,
        "hint": "", "required": "",
        "relevant": "", "constraint": "",
        "constraint_message": "", "appearance": "",
        "choice_filter": "", "calculation": "",
        "repeat_count": s.repeat_count or "", "parameters": "",
    }


def _section_close_row(s: FormSection) -> dict:
    closer = "end_repeat" if s.repeat_count else "end_group"
    return {
        "type": closer,
        "name": f"{s.name}_end",
        "label": "", "hint": "", "required": "",
        "relevant": "", "constraint": "",
        "constraint_message": "", "appearance": "",
        "choice_filter": "", "calculation": "",
        "repeat_count": "", "parameters": "",
    }


def _settings_row(form_version: FormVersion) -> tuple:
    """Kobo-shaped settings row. Mirrors the legacy file:
    `English` (not `en`), `theme-grid` (not `pages`), slugged
    form_id, datestamp version derived from updated_at so each
    re-export bumps the version even if FormVersion.version
    doesn't (Kobo uses version to detect "is this a new upload")."""
    form_id = f"{_slug(form_version.name)}_v{form_version.version}"
    version_stamp = form_version.updated_at.strftime("%Y%m%d%H%M")
    return (
        form_version.name,
        form_id,
        version_stamp,
        "English",
        "theme-grid",
    )


def export_to_xlsx(form_version: FormVersion) -> bytes:
    """Render `form_version` as a Kobo-compatible XLSForm. Returns
    raw .xlsx bytes — caller streams them as needed."""
    wb = Workbook()
    ws_survey = wb.active
    ws_survey.title = "survey"
    ws_choices = wb.create_sheet("choices")
    ws_settings = wb.create_sheet("settings")

    # ── survey sheet ────────────────────────────────────────────
    ws_survey.append(SURVEY_COLS)

    # Auto-prepend top-level metadata rows. The importer drops these
    # from FormQuestion (would otherwise duplicate them here on each
    # round-trip).
    for meta in _METADATA_ROWS:
        ws_survey.append([meta.get(c, "") for c in SURVEY_COLS])

    sections = list(
        form_version.sections.prefetch_related("questions__choice_list_ref")
        .order_by("order", "code"),
    )
    referenced_choice_lists: set[str] = set()
    referenced_geo_levels: set[str] = set()
    for sec in sections:
        ws_survey.append([_section_open_row(sec).get(c, "") for c in SURVEY_COLS])
        for q in sec.questions.order_by("order_in_section", "name"):
            # US-S21-006 — defensive skip: a calculate row with no
            # expression is what tripped Kobo with "[row : N]
            # Missing calculation." Skip it rather than ship an
            # invalid xlsx. The proper fix is to populate
            # FormQuestion.calculation; this safeguard prevents
            # one bad row from re-poisoning the whole upload.
            if q.type == "calculate" and not (q.calculation or ""):
                continue
            row = _question_row(q)
            ws_survey.append([row.get(c, "") for c in SURVEY_COLS])
            if q.name in GEO_QUESTIONS:
                referenced_geo_levels.add(GEO_QUESTIONS[q.name][0])
            elif q.choice_list_ref is not None:
                referenced_choice_lists.add(q.choice_list_ref.list_name)
        ws_survey.append([_section_close_row(sec).get(c, "") for c in SURVEY_COLS])

    # ── choices sheet ──────────────────────────────────────────
    # Only emit lists actually referenced — keeps the xlsx small.
    ws_choices.append(CHOICES_COLS)
    if referenced_choice_lists:
        from apps.reference_data.models import ChoiceList
        for cl in ChoiceList.objects.filter(
            list_name__in=referenced_choice_lists, version=1,
        ).order_by("list_name"):
            for opt in cl.options.filter(status="active").order_by("sort_order", "code"):
                ws_choices.append(
                    _padded_choice_row(cl.list_name, opt.code, opt.label),
                )
    # Geo lists — emit each level's options with the parent code in
    # the right ancestor column so Kobo's choice_filter resolves the
    # cascade at form runtime. Levels emitted in root→leaf order so
    # an operator opening the xlsx can read the cascade top-down.
    for level in GEO_LEVELS_ROOT_TO_LEAF:
        if level not in referenced_geo_levels:
            continue
        for opt in geo_options_for(level):
            ws_choices.append(
                _padded_choice_row(
                    level, opt["code"], opt["label"],
                    geo_parent_level=GEO_PARENT_COLUMN.get(level, ""),
                    geo_parent_code=opt.get("parent_code", ""),
                ),
            )

    # ── settings sheet ─────────────────────────────────────────
    ws_settings.append(SETTINGS_COLS)
    ws_settings.append(list(_settings_row(form_version)))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
