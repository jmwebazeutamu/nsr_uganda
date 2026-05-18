"""Intake submission tests."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest

from apps.data_management.models import Household
from apps.ingestion_hub.models import (
    Connector,
    DataProvisionAgreement,
    SourceSystem,
    SourceSystemKind,
    StageRecord,
    StageRecordState,
)
from apps.intake.models import (
    FormVersion,
    Submission,
    SubmissionResult,
    SubmissionState,
)
from apps.intake.services import IntakeError, submit_intake
from apps.reference_data.models import GeographicUnit

# --- Fixtures ---------------------------------------------------------------

@pytest.fixture
def geo(db):
    nodes = {}
    for level, key, parent in [
        ("region", "r", None), ("sub_region", "sr", "r"), ("district", "d", "sr"),
        ("county", "c", "d"), ("sub_county", "sc", "c"),
        ("parish", "p", "sc"), ("village", "v", "p"),
    ]:
        nodes[key] = GeographicUnit.objects.create(
            level=level, code=f"INT-{key.upper()}", name=key.title(),
            parent=nodes.get(parent), effective_from=date(2026, 1, 1),
        )
    return nodes


@pytest.fixture
def web_source_with_connector(db):
    src = SourceSystem.objects.create(code="WEB-OD", name="Web on-demand",
                                      kind=SourceSystemKind.WEB)
    DataProvisionAgreement.objects.create(
        source_system=src, reference="DPA-WEB-1",
        valid_from=date(2026, 1, 1), valid_to=date(2030, 12, 31),
    )
    return Connector.objects.create(source_system=src, name="web-default")


@pytest.fixture
def active_form_version(db):
    return FormVersion.objects.create(
        version=1, name="NSR Questionnaire", schema={"sections": ["id", "roster"]},
        is_active=True, effective_from=date(2026, 1, 1),
    )


def _payload(geo):
    return {
        "geographic": {
            "region": geo["r"].code, "sub_region": geo["sr"].code,
            "district": geo["d"].code, "county": geo["c"].code,
            "sub_county": geo["sc"].code, "parish": geo["p"].code,
            "village": geo["v"].code,
        },
        "urban_rural": "rural",
        "address_narrative": "Test homestead",
        "gps_lat": "1.234567", "gps_lng": "33.000000", "gps_accuracy_m": "5.00",
        "members": [
            {"line_number": 1, "surname": "Okot", "first_name": "James",
             "sex": "M", "relationship_to_head": "01", "is_head": True},
        ],
    }


# --- Refusal paths ----------------------------------------------------------

class TestSubmitIntakePreconditions:
    def test_no_active_form_version_raises(self, db, web_source_with_connector, geo):
        with pytest.raises(IntakeError, match="ACTIVE FormVersion"):
            submit_intake(
                channel="web", enumerator="e1",
                canonical_payload=_payload(geo),
            )

    def test_unsupported_channel_raises(self, db, active_form_version, geo):
        with pytest.raises(IntakeError, match="no DIH connector"):
            submit_intake(
                channel="partner_mis", enumerator="e1",
                canonical_payload=_payload(geo),
            )


# --- Happy path -------------------------------------------------------------

class TestSubmitIntake:
    def test_creates_submission_with_stage_and_provisional_id(
        self, db, web_source_with_connector, active_form_version, geo,
    ):
        sub = submit_intake(
            channel="web", enumerator="e1",
            canonical_payload=_payload(geo),
            auto_process=False,  # skip orchestrator for this test
        )
        assert sub.state == SubmissionState.PENDING_QA
        assert sub.result == SubmissionResult.COMPLETED
        assert sub.stage_record_id
        assert sub.provisional_registry_id
        # Stage row exists with the same provisional id.
        stage = StageRecord.objects.get(pk=sub.stage_record_id)
        assert stage.provisional_registry_id == sub.provisional_registry_id

    def test_auto_process_fast_tracks_clean_walkin_to_promoted(
        self, db, web_source_with_connector, active_form_version, geo,
    ):
        sub = submit_intake(
            channel="web", enumerator="e1",
            canonical_payload=_payload(geo),
            auto_process=True,
        )
        stage = StageRecord.objects.get(pk=sub.stage_record_id)
        # No DQA rules + no DDUP candidates + WEB channel -> AC-DIH-FT-AUTO.
        assert stage.state == StageRecordState.PROMOTED
        assert Household.objects.filter(pk=sub.provisional_registry_id).exists()


# --- HTTP surface -----------------------------------------------------------

class TestSubmitIntakeApi:
    def test_post_creates_submission_via_drf(
        self, db, web_source_with_connector, active_form_version, geo, django_user_model,
    ):
        from rest_framework.test import APIClient
        user = django_user_model.objects.create_user(
            username="enum-1", password="p", is_superuser=True, is_staff=True,
        )
        client = APIClient()
        client.force_authenticate(user=user)
        r = client.post(
            "/api/v1/intake/submissions/submit/",
            data={
                "channel": "web",
                "enumerator": "enum-1",
                "canonical_payload": _payload(geo),
                "auto_process": False,
            },
            format="json",
        )
        assert r.status_code == 200, r.content
        assert r.data["state"] == SubmissionState.PENDING_QA
        assert r.data["stage_record_id"]
        assert Submission.objects.filter(pk=r.data["id"]).exists()


# --- US-117a: questionnaire authoring models -------------------------------

class TestFormVersionLifecycleFields:
    """The new status/author/approval_note fields landed alongside the
    child-model structure. Defaults match the DqaRule/ChoiceList pattern."""

    def test_new_form_version_defaults_to_draft(self, db):
        fv = FormVersion.objects.create(
            version=999, name="test-form",
            effective_from=date(2026, 1, 1),
        )
        assert fv.status == "draft"
        assert fv.author == ""
        assert fv.approval_note == ""
        assert fv.submitted_at is None


class TestQuestionnaireSchema:
    @pytest.fixture
    def fv(self, db):
        return FormVersion.objects.create(
            version=2026, name="nsr-questionnaire",
            effective_from=date(2026, 1, 1),
            status="draft", author="alice",
        )

    def test_section_creation(self, fv):
        from apps.intake.models import FormSection
        s = FormSection.objects.create(
            form_version=fv, code="A", name="identification",
            label="Identification particulars", order=1,
        )
        assert str(s).endswith("A: Identification particulars")
        assert s.form_version_id == fv.id

    def test_section_code_unique_per_version(self, fv):
        from django.db import IntegrityError

        from apps.intake.models import FormSection
        FormSection.objects.create(
            form_version=fv, code="A", name="ident", label="x", order=1,
        )
        with pytest.raises(IntegrityError):
            FormSection.objects.create(
                form_version=fv, code="A", name="ident2", label="y", order=2,
            )

    def test_section_name_unique_per_version(self, db):
        from django.db import IntegrityError

        from apps.intake.models import FormSection
        fv = FormVersion.objects.create(
            version=2027, name="test-form",
            effective_from=date(2026, 1, 1),
        )
        FormSection.objects.create(
            form_version=fv, code="A", name="ident", label="x",
        )
        with pytest.raises(IntegrityError):
            FormSection.objects.create(
                form_version=fv, code="B", name="ident", label="y",
            )

    def test_question_with_choice_list_ref(self, fv):
        from apps.intake.models import FormQuestion, FormSection
        from apps.reference_data.models import ChoiceList
        # Use the seeded `relationship` choice list (v=1, status=active).
        rel = ChoiceList.objects.get(list_name="relationship", version=1)
        section = FormSection.objects.create(
            form_version=fv, code="C", name="roster", label="Household roster",
        )
        q = FormQuestion.objects.create(
            section=section, name="relationship_to_head",
            label="Relationship to head", type="select_one",
            choice_list_ref=rel, required=True, order_in_section=2,
        )
        assert q.choice_list_ref == rel
        assert q.type == "select_one"
        assert q.required is True

    def test_question_name_unique_per_section(self, fv):
        from django.db import IntegrityError

        from apps.intake.models import FormQuestion, FormSection
        s = FormSection.objects.create(
            form_version=fv, code="A", name="ident", label="x",
        )
        FormQuestion.objects.create(
            section=s, name="full_name", label="Full name", type="text",
        )
        with pytest.raises(IntegrityError):
            FormQuestion.objects.create(
                section=s, name="full_name", label="dup", type="text",
            )

    def test_skip_logic_and_constraint_attach_to_question(self, fv):
        from apps.intake.models import (
            FormConstraint,
            FormQuestion,
            FormSection,
            FormSkipLogic,
        )
        s = FormSection.objects.create(
            form_version=fv, code="C", name="roster", label="x",
        )
        q = FormQuestion.objects.create(
            section=s, name="age_years", label="Age in years", type="integer",
        )
        FormSkipLogic.objects.create(
            question=q,
            dsl={"field": "age_years", "op": "is_null"},
            description="don't ask when age missing",
        )
        FormConstraint.objects.create(
            question=q,
            dsl={"field": "age_years", "op": "between", "value": [0, 120]},
            message="age must be 0-120",
        )
        assert q.skip_logic.count() == 1
        assert q.constraints.count() == 1

    def test_section_questions_in_order(self, fv):
        from apps.intake.models import FormQuestion, FormSection
        s = FormSection.objects.create(
            form_version=fv, code="A", name="ident", label="x",
        )
        FormQuestion.objects.create(
            section=s, name="b", label="b", type="text", order_in_section=2,
        )
        FormQuestion.objects.create(
            section=s, name="a", label="a", type="text", order_in_section=1,
        )
        FormQuestion.objects.create(
            section=s, name="c", label="c", type="text", order_in_section=3,
        )
        names = list(s.questions.values_list("name", flat=True))
        assert names == ["a", "b", "c"]


class TestQuestionnaireAdminSmoke:
    """The admin pages render without 500 — the v2 tree UI lands in
    US-117b; this smoke test pins the inline + changelist contract."""

    def test_form_version_changelist(self, db, django_user_model):
        from django.test import Client
        # Need at least one row for the changelist to render column
        # headers — Django omits the table on an empty list.
        FormVersion.objects.create(
            version=2029, name="changelist-test", status="active",
            effective_from=date(2026, 1, 1),
        )
        u = django_user_model.objects.create_user(
            username="qa-staff", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        r = c.get("/admin/intake/formversion/")
        assert r.status_code == 200
        body = r.content.decode()
        # changelist-test row + Status column wired into list_display.
        assert "changelist-test" in body
        assert "Status" in body

    def test_form_section_admin(self, db, django_user_model):
        from django.test import Client

        from apps.intake.models import FormSection
        fv = FormVersion.objects.create(
            version=2028, name="admin-test-form",
            effective_from=date(2026, 1, 1),
        )
        FormSection.objects.create(
            form_version=fv, code="A", name="ident",
            label="Identification", order=1,
        )
        u = django_user_model.objects.create_user(
            username="qa-staff2", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        r = c.get("/admin/intake/formsection/")
        assert r.status_code == 200
        assert "Identification" in r.content.decode()


# --- US-117b: builder UI (tree + reorder + validate) -----------------------

class TestQuestionnaireBuilderUI:
    """The change_form template renders the tree pane + the
    expression-validator panel when QUESTIONNAIRE_EDITOR_V2 is on.
    Falls back to the default admin form when off."""

    def _staff_client(self, db, django_user_model):
        from django.test import Client
        u = django_user_model.objects.create_user(
            username="qb-staff", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        return c

    def _seeded_fv(self, db):
        from apps.intake.models import FormQuestion, FormSection, FormVersion
        fv = FormVersion.objects.create(
            version=3000, name="qb-test",
            effective_from=date(2026, 1, 1),
            status="draft", author="alice",
        )
        a = FormSection.objects.create(
            form_version=fv, code="A", name="ident", label="Identification", order=1,
        )
        b = FormSection.objects.create(
            form_version=fv, code="B", name="status", label="Survey status", order=2,
        )
        q1 = FormQuestion.objects.create(
            section=a, name="full_name", label="Full name",
            type="text", required=True, order_in_section=1,
        )
        q2 = FormQuestion.objects.create(
            section=a, name="phone", label="Phone",
            type="text", required=False, order_in_section=2,
        )
        return fv, a, b, q1, q2

    def test_change_form_renders_tree_when_flag_on(
        self, db, django_user_model, settings,
    ):
        settings.QUESTIONNAIRE_EDITOR_V2 = True
        fv, a, b, q1, q2 = self._seeded_fv(db)
        r = self._staff_client(db, django_user_model).get(
            f"/admin/intake/formversion/{fv.id}/change/",
        )
        assert r.status_code == 200
        body = r.content.decode()
        # Tree pane present.
        assert 'id="qe-tree"' in body
        # Both sections rendered (by code).
        assert "Identification" in body and "Survey status" in body
        # Both questions rendered (by name).
        assert "full_name" in body and "phone" in body
        # Validator panel present.
        assert 'id="qe-validate"' in body

    def test_change_form_omits_tree_when_flag_off(
        self, db, django_user_model, settings,
    ):
        settings.QUESTIONNAIRE_EDITOR_V2 = False
        fv, *_ = self._seeded_fv(db)
        r = self._staff_client(db, django_user_model).get(
            f"/admin/intake/formversion/{fv.id}/change/",
        )
        assert r.status_code == 200
        body = r.content.decode()
        assert 'id="qe-tree"' not in body
        assert 'id="qe-validate"' not in body

    def test_reorder_section_swaps_order(self, db, django_user_model):
        import json as _json
        fv, a, b, *_ = self._seeded_fv(db)
        # Initially: a.order=1, b.order=2. Move b up → swap.
        c = self._staff_client(db, django_user_model)
        r = c.post(
            f"/admin/intake/formversion/_us117b/reorder-section/{b.id}/",
            data=_json.dumps({"direction": "up"}),
            content_type="application/json",
        )
        assert r.status_code == 200
        a.refresh_from_db()
        b.refresh_from_db()
        assert b.order < a.order

    def test_reorder_section_boundary_no_op(self, db, django_user_model):
        import json as _json
        fv, a, *_ = self._seeded_fv(db)
        c = self._staff_client(db, django_user_model)
        r = c.post(
            f"/admin/intake/formversion/_us117b/reorder-section/{a.id}/",
            data=_json.dumps({"direction": "up"}),
            content_type="application/json",
        )
        assert r.status_code == 200
        body = r.json()
        assert body["moved"] is False

    def test_reorder_question_swaps_within_section(self, db, django_user_model):
        import json as _json
        fv, a, b, q1, q2 = self._seeded_fv(db)
        c = self._staff_client(db, django_user_model)
        # q2 (order=2) moves up over q1 (order=1).
        r = c.post(
            f"/admin/intake/formversion/_us117b/reorder-question/{q2.id}/",
            data=_json.dumps({"direction": "up"}),
            content_type="application/json",
        )
        assert r.status_code == 200
        q1.refresh_from_db()
        q2.refresh_from_db()
        assert q2.order_in_section < q1.order_in_section

    # --- US-117c: drag-and-drop "after_id" payload ----

    def test_reorder_section_after_id_drops_below_target(
        self, db, django_user_model,
    ):
        """Drag section a (order=1) and drop it after section b
        (order=2) → resulting order [b, a]."""
        import json as _json

        from apps.intake.models import FormSection
        fv, a, b, *_ = self._seeded_fv(db)
        c = self._staff_client(db, django_user_model)
        r = c.post(
            f"/admin/intake/formversion/_us117b/reorder-section/{a.id}/",
            data=_json.dumps({"after_id": b.id}),
            content_type="application/json",
        )
        assert r.status_code == 200
        assert r.json()["moved"] is True
        ordered = list(
            FormSection.objects.filter(form_version=fv)
            .order_by("order").values_list("id", flat=True),
        )
        assert ordered == [b.id, a.id]

    def test_reorder_section_after_id_empty_lands_at_top(
        self, db, django_user_model,
    ):
        """An empty after_id ("") puts the section at position 0."""
        import json as _json

        from apps.intake.models import FormSection
        fv, a, b, *_ = self._seeded_fv(db)
        c = self._staff_client(db, django_user_model)
        # Move b to the top.
        r = c.post(
            f"/admin/intake/formversion/_us117b/reorder-section/{b.id}/",
            data=_json.dumps({"after_id": ""}),
            content_type="application/json",
        )
        assert r.status_code == 200
        ordered = list(
            FormSection.objects.filter(form_version=fv)
            .order_by("order").values_list("id", flat=True),
        )
        assert ordered == [b.id, a.id]

    def test_reorder_question_after_id_drops_below_target(
        self, db, django_user_model,
    ):
        """Drag q1 (order=1) after q2 (order=2) → resulting order [q2, q1]."""
        import json as _json

        from apps.intake.models import FormQuestion
        fv, a, b, q1, q2 = self._seeded_fv(db)
        c = self._staff_client(db, django_user_model)
        r = c.post(
            f"/admin/intake/formversion/_us117b/reorder-question/{q1.id}/",
            data=_json.dumps({"after_id": q2.id}),
            content_type="application/json",
        )
        assert r.status_code == 200
        ordered = list(
            FormQuestion.objects.filter(section=a)
            .order_by("order_in_section").values_list("id", flat=True),
        )
        assert ordered == [q2.id, q1.id]

    def test_reorder_question_unknown_after_id_returns_no_op(
        self, db, django_user_model,
    ):
        import json as _json
        fv, a, b, q1, q2 = self._seeded_fv(db)
        c = self._staff_client(db, django_user_model)
        r = c.post(
            f"/admin/intake/formversion/_us117b/reorder-question/{q1.id}/",
            data=_json.dumps({"after_id": "no-such-id-XYZ"}),
            content_type="application/json",
        )
        assert r.status_code == 200
        body = r.json()
        assert body["moved"] is False

    def test_change_form_tree_marks_rows_draggable(
        self, db, django_user_model, settings,
    ):
        """The HTML5 drag-and-drop relies on draggable=true on each
        section and question row. The CSS contract is tested by the
        markup containing the affordance hint."""
        settings.QUESTIONNAIRE_EDITOR_V2 = True
        fv, *_ = self._seeded_fv(db)
        r = self._staff_client(db, django_user_model).get(
            f"/admin/intake/formversion/{fv.id}/change/",
        )
        body = r.content.decode()
        assert "qe-drag-handle" in body
        assert 'draggable="true"' in body
        assert "drag rows to reorder" in body

    # --- US-117c-2: cross-section question drag ----

    def test_reorder_question_can_cross_sections_via_after_id(
        self, db, django_user_model,
    ):
        """Drag q1 (in section A) and drop after a question in
        section B → q1 re-parents to section B and lands after the
        anchor."""
        import json as _json

        from apps.intake.models import FormQuestion
        fv, a, b, q1, q2 = self._seeded_fv(db)
        q3 = FormQuestion.objects.create(
            section=b, name="b_anchor", label="Anchor",
            type="text", order_in_section=1,
        )
        c = self._staff_client(db, django_user_model)
        r = c.post(
            f"/admin/intake/formversion/_us117b/reorder-question/{q1.id}/",
            data=_json.dumps({"after_id": q3.id}),
            content_type="application/json",
        )
        assert r.status_code == 200
        body = r.json()
        assert body["moved"] is True
        assert body["re_parented"] is True
        b_order = list(
            FormQuestion.objects.filter(section=b)
            .order_by("order_in_section").values_list("id", flat=True),
        )
        a_order = list(
            FormQuestion.objects.filter(section=a)
            .order_by("order_in_section").values_list("id", flat=True),
        )
        assert b_order == [q3.id, q1.id]
        assert a_order == [q2.id]

    def test_reorder_question_explicit_section_id_top_of_section(
        self, db, django_user_model,
    ):
        """Explicit section_id with after_id="" lands the question
        at position 0 of the target section — used when the operator
        drops onto a section header (populating an empty section)."""
        import json as _json

        from apps.intake.models import FormQuestion
        fv, a, b, q1, q2 = self._seeded_fv(db)
        c = self._staff_client(db, django_user_model)
        r = c.post(
            f"/admin/intake/formversion/_us117b/reorder-question/{q1.id}/",
            data=_json.dumps({"section_id": b.id, "after_id": ""}),
            content_type="application/json",
        )
        assert r.status_code == 200
        b_order = list(
            FormQuestion.objects.filter(section=b)
            .order_by("order_in_section").values_list("id", flat=True),
        )
        assert b_order == [q1.id]

    def test_reorder_question_rejects_cross_form_version(
        self, db, django_user_model,
    ):
        """Question can move between sections of the same FormVersion
        but not into a section of a different FormVersion — the
        meaning would change under past Submissions."""
        import json as _json

        from apps.intake.models import FormSection, FormVersion
        fv, a, b, q1, q2 = self._seeded_fv(db)
        other_fv = FormVersion.objects.create(
            version=3001, name="other", effective_from=date(2026, 1, 1),
            status="draft", author="qa",
        )
        other_sec = FormSection.objects.create(
            form_version=other_fv, code="X", name="x", label="X", order=1,
        )
        c = self._staff_client(db, django_user_model)
        r = c.post(
            f"/admin/intake/formversion/_us117b/reorder-question/{q1.id}/",
            data=_json.dumps({"section_id": other_sec.id, "after_id": ""}),
            content_type="application/json",
        )
        assert r.status_code == 400
        assert "cannot move question across FormVersions" in r.json()["detail"]

    def test_validate_expression_ok(self, db, django_user_model):
        import json as _json
        c = self._staff_client(db, django_user_model)
        r = c.post(
            "/admin/intake/formversion/_us117b/validate-expression/",
            data=_json.dumps({
                "expression": {"field": "age", "op": "between",
                               "value": [0, 120]},
                "sample_record": {"age": 25},
            }),
            content_type="application/json",
        )
        assert r.status_code == 200
        body = r.json()
        assert body["ok"] is True
        assert body["result"] is True

    def test_validate_expression_dsl_error(self, db, django_user_model):
        import json as _json
        c = self._staff_client(db, django_user_model)
        r = c.post(
            "/admin/intake/formversion/_us117b/validate-expression/",
            data=_json.dumps({
                "expression": {"op": "no_such_op", "field": "x"},
                "sample_record": {},
            }),
            content_type="application/json",
        )
        assert r.status_code == 200
        body = r.json()
        assert body["ok"] is False
        assert "no_such_op" in body["error"]

    def test_validate_rejects_non_dict_expression(self, db, django_user_model):
        import json as _json
        c = self._staff_client(db, django_user_model)
        r = c.post(
            "/admin/intake/formversion/_us117b/validate-expression/",
            data=_json.dumps({"expression": "not-a-dict"}),
            content_type="application/json",
        )
        assert r.status_code == 400


# --- US-120: legacy questionnaire import -----------------------------------

# The legacy builder lives in /k-forms/build_nsr_xlsform.py and is not
# committed to the repo (it pulls a personal OneDrive geo workbook
# path). Without it, the importer can't run, so the whole class is
# skipped in CI where the directory is absent. Local dev runs against
# the file as it sits in the developer's working tree.
_LEGACY_SCRIPT = (
    Path(__file__).resolve().parent.parent.parent
    / "k-forms" / "build_nsr_xlsform.py"
)
_skip_without_legacy = pytest.mark.skipif(
    not _LEGACY_SCRIPT.exists(),
    reason=(
        "k-forms/build_nsr_xlsform.py is not present — it lives outside "
        "the repo by design (hard-coded developer paths). Local dev has it."
    ),
)


@_skip_without_legacy
class TestLegacyQuestionnaireImport:
    """The scripts/import_legacy_questionnaire.py end-to-end check.

    Reads k-forms/build_nsr_xlsform.py via exec (Workbook.save
    monkey-patched no-op) and builds FormVersion v1. Re-running
    rebuilds the children in place — no stale rows survive.
    """

    @pytest.fixture
    def _seed_choice_lists(self, db):
        """Migration-applied seed isn't visible in TransactionTestCase-style
        tests — call into the migration data loader directly here so the
        importer's ChoiceList look-ups resolve."""
        import json
        from datetime import date as _d
        from pathlib import Path

        from apps.reference_data.models import ChoiceList, ChoiceOption
        path = (
            Path(__file__).resolve().parent.parent / "reference_data"
            / "seeds" / "choice_lists_v1.json"
        )
        for name, options in json.loads(path.read_text()).items():
            cl, _ = ChoiceList.objects.get_or_create(
                list_name=name, version=1,
                defaults={"effective_from": _d(2026, 1, 1),
                          "status": "active",
                          "author": "system-migration",
                          "approved_by": "system-migration"},
            )
            for so, opt in enumerate(options, start=1):
                ChoiceOption.objects.get_or_create(
                    choice_list=cl, code=opt["code"], language="en",
                    defaults={"label": opt["label"], "sort_order": so},
                )

    def test_import_creates_form_version_one(self, _seed_choice_lists):
        from scripts.import_legacy_questionnaire import main

        from apps.intake.models import FormVersion
        result = main()
        fv = FormVersion.objects.get(version=1)
        assert fv.id == result["form_version_id"]
        assert fv.status == "active"
        assert fv.is_active is True
        # At least the 9 top-level sections we know about land.
        assert fv.sections.count() >= 9

    def test_import_links_choice_list_refs(self, _seed_choice_lists):
        from scripts.import_legacy_questionnaire import main

        from apps.intake.models import FormQuestion
        main()
        # The legacy script names the C2 column `c2_relationship`.
        q = FormQuestion.objects.filter(name="c2_relationship").first()
        assert q is not None, "c2_relationship question not imported"
        assert q.choice_list_ref is not None
        assert q.choice_list_ref.list_name == "relationship"
        # And confirm at least one select_one question landed with a
        # resolved ChoiceList ref overall.
        with_refs = FormQuestion.objects.filter(
            choice_list_ref__isnull=False,
        ).count()
        assert with_refs >= 20  # legacy form has dozens of selects

    def test_import_is_idempotent(self, _seed_choice_lists):
        from scripts.import_legacy_questionnaire import main

        from apps.intake.models import FormQuestion, FormSection
        main()
        sections_before = FormSection.objects.count()
        questions_before = FormQuestion.objects.count()
        main()
        # Re-run rebuilds in place — counts stay stable, no orphan
        # rows from the prior import.
        assert FormSection.objects.count() == sections_before
        assert FormQuestion.objects.count() == questions_before

    def test_import_preserves_xlsform_metadata_types(self, _seed_choice_lists):
        """US-120b regression: an earlier importer mapped any unknown
        legacy type (notably `time`) to END_GROUP via a misguided
        fallback. The `a15_start_time` row was the canary."""
        from scripts.import_legacy_questionnaire import main

        from apps.intake.models import FormQuestion
        main()
        q = FormQuestion.objects.get(name="a15_start_time")
        assert q.type == "time", (
            f"a15_start_time should import as type=time, got {q.type!r} — "
            "QuestionType enum is likely missing `time` again."
        )

    def test_import_attaches_pre_section_questions_to_prior_section(
        self, _seed_choice_lists,
    ):
        """US-120b regression: `hh_size` sits at the top level in the
        legacy script (between sections B and C) and was dropped on
        import. It now attaches to the most-recently-closed section
        so the begin_repeat that follows can reference it."""
        from scripts.import_legacy_questionnaire import main

        from apps.intake.models import FormQuestion
        main()
        q = FormQuestion.objects.get(name="hh_size")
        assert q.type == "integer"
        assert q.section.name == "survey_status"

    def test_import_marks_household_roster_as_repeat(self, _seed_choice_lists):
        """US-120b regression: the household_members section is a
        begin_repeat in the legacy script. The importer must capture
        its repeat_count attribute on the FormSection so the export
        round-trip re-emits a begin_repeat (not begin_group)."""
        from scripts.import_legacy_questionnaire import main

        from apps.intake.models import FormSection
        main()
        sec = FormSection.objects.get(name="household_members")
        assert sec.repeat_count == "${hh_size}", (
            f"household_members.repeat_count should be ${{hh_size}}, "
            f"got {sec.repeat_count!r}"
        )


# --- US-118: XLSForm export from FormVersion -------------------------------

@_skip_without_legacy
class TestXlsformExport:
    """Round-trip: import legacy script → export → bytes are a valid
    XLSForm with the expected sheets/columns/rows. Tests pin the
    structural contract; not every row is asserted."""

    @pytest.fixture
    def _seeded_form(self, db):
        import json
        from datetime import date as _d
        from pathlib import Path

        from scripts.import_legacy_questionnaire import main as import_main

        from apps.intake.models import FormVersion
        from apps.reference_data.models import ChoiceList, ChoiceOption
        path = (
            Path(__file__).resolve().parent.parent / "reference_data"
            / "seeds" / "choice_lists_v1.json"
        )
        for name, options in json.loads(path.read_text()).items():
            cl, _ = ChoiceList.objects.get_or_create(
                list_name=name, version=1,
                defaults={"effective_from": _d(2026, 1, 1),
                          "status": "active",
                          "author": "system-migration",
                          "approved_by": "system-migration"},
            )
            for so, opt in enumerate(options, start=1):
                ChoiceOption.objects.get_or_create(
                    choice_list=cl, code=opt["code"], language="en",
                    defaults={"label": opt["label"], "sort_order": so},
                )
        import_main()
        return FormVersion.objects.get(version=1)

    def test_export_returns_xlsx_bytes(self, _seeded_form):
        from apps.intake.xlsform_export import export_to_xlsx
        out = export_to_xlsx(_seeded_form)
        assert isinstance(out, bytes)
        assert out[:2] == b"PK"  # ZIP magic — .xlsx is a zip

    def test_export_has_required_sheets(self, _seeded_form):
        import io

        import openpyxl

        from apps.intake.xlsform_export import export_to_xlsx
        wb = openpyxl.load_workbook(io.BytesIO(export_to_xlsx(_seeded_form)))
        assert {"survey", "choices", "settings"}.issubset(set(wb.sheetnames))

    def test_export_survey_sheet_has_questions(self, _seeded_form):
        import io

        import openpyxl

        from apps.intake.xlsform_export import export_to_xlsx
        wb = openpyxl.load_workbook(io.BytesIO(export_to_xlsx(_seeded_form)))
        ws = wb["survey"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) >= 100  # 184 questions + structural markers
        types = {r[0] for r in rows if r and r[0]}
        assert "begin_group" in types
        assert any(t and t.startswith("select_one") for t in types)

    def test_export_choices_sheet_has_options(self, _seeded_form):
        import io

        import openpyxl

        from apps.intake.xlsform_export import export_to_xlsx
        wb = openpyxl.load_workbook(io.BytesIO(export_to_xlsx(_seeded_form)))
        ws = wb["choices"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) >= 100
        names = {r[0] for r in rows if r and r[0]}
        assert "relationship" in names and "marital_status" in names

    def test_export_settings_sheet_has_form_metadata(self, _seeded_form):
        import io

        import openpyxl

        from apps.intake.xlsform_export import export_to_xlsx
        wb = openpyxl.load_workbook(io.BytesIO(export_to_xlsx(_seeded_form)))
        ws = wb["settings"]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        assert len(rows) >= 2
        header = {h for h in rows[0] if h}
        assert {"form_title", "form_id", "version"}.issubset(header)

    def test_admin_export_action_returns_xlsx(
        self, _seeded_form, db, django_user_model,
    ):
        from django.test import Client
        u = django_user_model.objects.create_user(
            username="xform-staff", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        r = c.get(
            f"/admin/intake/formversion/_us118/export-xlsform/{_seeded_form.id}/",
        )
        assert r.status_code == 200
        assert "spreadsheetml" in r["Content-Type"]
        assert r.content[:2] == b"PK"

    def test_export_settings_uses_kobo_valid_values(self, _seeded_form):
        """US-118b regression: Kobo's web preview rejects forms whose
        settings sheet uses `default_language=en` (it expects the
        full language name) or `style=pages` (legacy used theme-grid).
        Also pins the form_id slug + datestamp version format."""
        import io

        import openpyxl

        from apps.intake.xlsform_export import export_to_xlsx
        wb = openpyxl.load_workbook(io.BytesIO(export_to_xlsx(_seeded_form)))
        ws = wb["settings"]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        header = list(rows[0])
        values = dict(zip(header, rows[1], strict=False))
        assert values["default_language"] == "English"
        assert values["style"] == "theme-grid"
        # form_id should be a slug, not the literal FormVersion.name
        # (which contains spaces).
        assert " " not in values["form_id"]
        # Version stamp is YYYYMMDDHHMM, not just the int version.
        assert len(values["version"]) == 12 and values["version"].isdigit()

    def test_export_prepends_top_level_metadata_rows(self, _seeded_form):
        """US-118b regression: every XLSForm needs start/end/today/
        deviceid/username rows at the top so Kobo can stamp submission
        metadata. The legacy file had them; ours didn't."""
        import io

        import openpyxl

        from apps.intake.xlsform_export import export_to_xlsx
        wb = openpyxl.load_workbook(io.BytesIO(export_to_xlsx(_seeded_form)))
        ws = wb["survey"]
        rows = list(ws.iter_rows(min_row=2, max_row=6, values_only=True))
        types = [r[0] for r in rows]
        assert types == ["start", "end", "today", "deviceid", "username"]

    def test_export_emits_begin_repeat_for_roster_section(self, _seeded_form):
        """US-118b regression: household_members is a roster — exported
        as begin_repeat with repeat_count, not begin_group. Without
        this, every form rendered in Kobo collapses the roster into a
        non-iterable group and only one member can be entered."""
        import io

        import openpyxl

        from apps.intake.xlsform_export import export_to_xlsx
        wb = openpyxl.load_workbook(io.BytesIO(export_to_xlsx(_seeded_form)))
        rows = list(wb["survey"].iter_rows(min_row=1, values_only=True))
        header = rows[0]
        type_col = header.index("type")
        name_col = header.index("name")
        repeat_col = header.index("repeat_count")
        roster_open = next(
            r for r in rows[1:]
            if r[type_col] == "begin_repeat" and r[name_col] == "household_members"
        )
        assert roster_open[repeat_col] == "${hh_size}"
        # And its closer is end_repeat (not end_group).
        roster_close = next(
            r for r in rows[1:]
            if r[type_col] == "end_repeat" and r[name_col] == "household_members_end"
        )
        assert roster_close is not None

    def test_export_emits_geo_selects_with_cascading_filter(
        self, _seeded_form,
    ):
        """US-S20-006: geo selects (a0_region through a5_parish_ward)
        now emit as real select_one rows with proper list names and a
        choice_filter referencing the parent question — no more `text`
        fallback in the export. Verifies the cascade reaches Kobo."""
        import io

        import openpyxl

        from apps.intake.xlsform_export import export_to_xlsx
        wb = openpyxl.load_workbook(io.BytesIO(export_to_xlsx(_seeded_form)))
        rows = list(wb["survey"].iter_rows(min_row=1, values_only=True))
        header = rows[0]
        type_col = header.index("type")
        name_col = header.index("name")
        choice_filter_col = header.index("choice_filter")
        appearance_col = header.index("appearance")

        region_row = next(r for r in rows[1:] if r[name_col] == "a0_region")
        assert region_row[type_col] == "select_one region"
        # Top-level — no choice_filter.
        assert (region_row[choice_filter_col] or "") == ""
        # Geo questions inherit appearance=minimal when blank.
        assert region_row[appearance_col] == "minimal"

        subregion_row = next(r for r in rows[1:] if r[name_col] == "a1_subregion")
        assert subregion_row[type_col] == "select_one sub_region"
        assert subregion_row[choice_filter_col] == "region=${a0_region}"

        parish_row = next(r for r in rows[1:] if r[name_col] == "a5_parish_ward")
        assert parish_row[type_col] == "select_one parish"
        assert parish_row[choice_filter_col] == "sub_county=${a4_subcounty_division_tc}"

        # No bare select_one rows leak through.
        for r in rows[1:]:
            assert r[type_col] not in ("select_one", "select_multiple"), (
                f"row {r[name_col]} has bare {r[type_col]!r} — Kobo will reject"
            )

    def test_export_choices_sheet_carries_geo_ancestor_columns(
        self, _seeded_form,
    ):
        """The cascade only works if the choices sheet has ancestor
        columns populated — sub_region rows need a `region` cell so
        Kobo can apply `region=${a0_region}`. This test validates the
        wiring from REF-DATA.GeographicUnit through to the xlsx."""
        import io

        import openpyxl

        from apps.intake.xlsform_export import export_to_xlsx
        from apps.reference_data.models import GeographicUnit

        # Seed a known geo pair the test can pin against.
        r1 = GeographicUnit.objects.create(
            level="region", code="TR-1", name="TestRegion",
            effective_from=date(2026, 1, 1), status="active",
        )
        GeographicUnit.objects.create(
            level="sub_region", code="TSR-1", name="TestSubregion",
            parent=r1,
            effective_from=date(2026, 1, 1), status="active",
        )
        wb = openpyxl.load_workbook(io.BytesIO(export_to_xlsx(_seeded_form)))
        rows = list(wb["choices"].iter_rows(min_row=1, values_only=True))
        header = list(rows[0])
        assert "region" in header and "sub_region" in header
        list_col = header.index("list_name")
        name_col = header.index("name")
        region_parent_col = header.index("region")

        # Find the sub_region row for our seeded TSR-1 and confirm
        # its `region` cell carries the parent's code.
        sub_row = next(
            r for r in rows[1:]
            if r[list_col] == "sub_region" and r[name_col] == "TSR-1"
        )
        assert sub_row[region_parent_col] == "TR-1"

        # And the top-level region row leaves all ancestor columns blank.
        region_row = next(
            r for r in rows[1:]
            if r[list_col] == "region" and r[name_col] == "TR-1"
        )
        assert (region_row[region_parent_col] or "") == ""

    def test_calculate_row_carries_expression(self, _seeded_form):
        """US-S21-006 regression — Kobo rejected the legacy roster
        with '[row : 38] Missing calculation.' because FormQuestion
        had no `calculation` field, so the importer dropped
        `position(..)` and the exporter wrote an empty cell.
        Verifies the round-trip now ships the expression."""
        import io

        import openpyxl

        from apps.intake.xlsform_export import export_to_xlsx
        wb = openpyxl.load_workbook(io.BytesIO(export_to_xlsx(_seeded_form)))
        rows = list(wb["survey"].iter_rows(min_row=1, values_only=True))
        header = list(rows[0])
        type_col = header.index("type")
        name_col = header.index("name")
        calc_col = header.index("calculation")
        # Every calculate row must have a non-empty calculation —
        # otherwise Kobo rejects the deploy.
        calc_rows = [r for r in rows[1:] if r[type_col] == "calculate"]
        assert calc_rows, "expected at least one calculate row in the v1 export"
        for r in calc_rows:
            assert (r[calc_col] or "").strip(), (
                f"calculate row name={r[name_col]!r} has empty "
                "calculation — Kobo will reject this xlsx"
            )
        # member_index specifically carries position(..) per the legacy script.
        member_index = next(
            r for r in calc_rows if r[name_col] == "member_index"
        )
        assert member_index[calc_col] == "position(..)"

    def test_export_skips_orphan_calculate_with_no_expression(self, db):
        """Defensive: if a calculate row sneaks in with an empty
        expression (data migration, manual admin entry, broken
        importer), the exporter drops it from the survey rather
        than ship an unloadable xlsx."""
        import io

        import openpyxl

        from apps.intake.models import FormQuestion, FormSection, FormVersion
        from apps.intake.xlsform_export import export_to_xlsx
        fv = FormVersion.objects.create(
            version=8200, name="orphan-calc",
            effective_from=date(2026, 1, 1),
            status="draft", author="qa",
        )
        s = FormSection.objects.create(
            form_version=fv, code="A", name="ident", label="A", order=1,
        )
        FormQuestion.objects.create(
            section=s, name="ok_q", label="OK", type="text",
            order_in_section=1,
        )
        FormQuestion.objects.create(
            section=s, name="orphan_calc", label="orphan",
            type="calculate", order_in_section=2,
            # calculation deliberately blank
        )
        wb = openpyxl.load_workbook(io.BytesIO(export_to_xlsx(fv)))
        names = [r[1] for r in wb["survey"].iter_rows(min_row=2, values_only=True)]
        assert "ok_q" in names
        assert "orphan_calc" not in names  # silently skipped


# --- US-117e: interactive in-admin preview ---------------------------------

class TestInteractivePreview:
    """US-117e ships an interactive preview at
    /admin/intake/formversion/<id>/_us117e/preview/ that hydrates a
    React harness from a JSON schema embedded in the page. Skip-logic,
    constraints, and repeat-section (roster) add/remove all fire
    client-side. Tests here pin the server contract — the schema
    shape, the URL surface, and the staff-gate."""

    @pytest.fixture
    def _fv_with_roster(self, db):
        from apps.intake.models import (
            FormQuestion,
            FormSection,
            FormVersion,
        )
        from apps.reference_data.models import ChoiceList, ChoiceOption
        fv = FormVersion.objects.create(
            version=5117, name="interactive-preview-fixture",
            effective_from=date(2026, 1, 1),
            status="draft", author="qa",
        )
        # A consent section that gates the rest of the form.
        s_consent = FormSection.objects.create(
            form_version=fv, code="CONSENT", name="consent_group",
            label="Consent", order=1,
        )
        cl_yn = ChoiceList.objects.create(
            list_name="ip_yes_no", version=1,
            effective_from=date(2026, 1, 1),
            status="active",
            author="qa", approved_by="qa",
        )
        ChoiceOption.objects.create(
            choice_list=cl_yn, code="1", label="Yes",
            language="en", sort_order=1,
        )
        ChoiceOption.objects.create(
            choice_list=cl_yn, code="2", label="No",
            language="en", sort_order=2,
        )
        FormQuestion.objects.create(
            section=s_consent, name="consent", label="Consent?",
            type="select_one", choice_list_ref=cl_yn,
            required=True, order_in_section=1,
        )
        # A roster section with a per-member age question.
        s_roster = FormSection.objects.create(
            form_version=fv, code="C", name="household_members",
            label="Household roster", order=2,
            repeat_count="${hh_size}",
        )
        FormQuestion.objects.create(
            section=s_roster, name="age_years", label="Age",
            type="integer", order_in_section=1,
            constraint_expression=". >= 0 and . <= 120",
            constraint_message="Age must be 0-120",
            relevant_expression="${consent}='1'",
        )
        return fv

    def test_schema_endpoint_returns_form_shape(
        self, _fv_with_roster, django_user_model,
    ):
        from django.test import Client
        u = django_user_model.objects.create_user(
            username="ip-staff", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        r = c.get(
            f"/admin/intake/formversion/_us117e/schema/{_fv_with_roster.id}/",
        )
        assert r.status_code == 200
        body = r.json()
        assert body["name"] == "interactive-preview-fixture"
        assert body["version"] == 5117
        # Sections come through in order with their roster flag.
        names = [s["name"] for s in body["sections"]]
        assert names == ["consent_group", "household_members"]
        roster = body["sections"][1]
        assert roster["repeat_count"] == "${hh_size}"
        # Inlined choice options on the consent question.
        consent_q = body["sections"][0]["questions"][0]
        assert consent_q["type"] == "select_one"
        assert {o["code"] for o in consent_q["options"]} == {"1", "2"}
        # Constraint + relevant expressions ride through verbatim.
        age_q = roster["questions"][0]
        assert age_q["constraint"] == ". >= 0 and . <= 120"
        assert age_q["relevant"] == "${consent}='1'"

    def test_interactive_preview_view_embeds_schema(
        self, _fv_with_roster, django_user_model,
    ):
        from django.test import Client
        u = django_user_model.objects.create_user(
            username="ip-staff2", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        r = c.get(
            f"/admin/intake/formversion/_us117e/preview/{_fv_with_roster.id}/",
        )
        assert r.status_code == 200
        body = r.content.decode("utf-8")
        # The schema is embedded via json_script — assert both the
        # script tag id and a sentinel from the schema body.
        assert 'id="us117e-schema"' in body
        assert "household_members" in body
        # The React mount point lives in the template.
        assert 'id="us117e-root"' in body
        # And the action bar links back to the static preview and
        # XLSForm download.
        assert f"/admin/intake/formversion/_us117b/preview/{_fv_with_roster.id}/" in body
        assert f"/admin/intake/formversion/_us118/export-xlsform/{_fv_with_roster.id}/" in body

    def test_preview_resolves_by_version_number(
        self, _fv_with_roster, django_user_model,
    ):
        """Mirror US-117d's `_resolve_form_version` convenience —
        an admin should be able to type the version number into
        the URL bar instead of the ULID."""
        from django.test import Client
        u = django_user_model.objects.create_user(
            username="ip-staff3", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        r = c.get(
            f"/admin/intake/formversion/_us117e/preview/{_fv_with_roster.version}/",
        )
        assert r.status_code == 200

    def test_preview_404_on_unknown(self, db, django_user_model):
        from django.test import Client
        u = django_user_model.objects.create_user(
            username="ip-staff4", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        r = c.get("/admin/intake/formversion/_us117e/preview/999999/")
        assert r.status_code == 404
        r = c.get("/admin/intake/formversion/_us117e/schema/999999/")
        assert r.status_code == 404

    def test_schema_inlines_geo_options_with_parent_code(
        self, db, django_user_model,
    ):
        """The 6 legacy geographic select questions have no
        ChoiceList in the seed catalogue — schema instead inlines
        options from REF-DATA.GeographicUnit with parent_code so
        React can cascade-filter (region → subregion → … → parish)."""
        from django.test import Client

        from apps.intake.models import FormQuestion, FormSection, FormVersion
        from apps.reference_data.models import GeographicUnit
        # Two-level synthetic hierarchy in the test DB.
        ug = GeographicUnit.objects.create(
            level="region", code="R-TEST", name="Test Region",
            effective_from=date(2026, 1, 1), status="active",
        )
        GeographicUnit.objects.create(
            level="sub_region", code="SR-TEST", name="Test Subregion",
            parent=ug,
            effective_from=date(2026, 1, 1), status="active",
        )
        fv = FormVersion.objects.create(
            version=5118, name="geo-preview-fixture",
            effective_from=date(2026, 1, 1),
            status="draft", author="qa",
        )
        s = FormSection.objects.create(
            form_version=fv, code="A", name="identification",
            label="Identification", order=1,
        )
        FormQuestion.objects.create(
            section=s, name="a0_region", label="Region",
            type="select_one", required=True, order_in_section=1,
        )
        FormQuestion.objects.create(
            section=s, name="a1_subregion", label="Subregion",
            type="select_one", required=True, order_in_section=2,
        )
        u = django_user_model.objects.create_user(
            username="ip-staff-geo", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        r = c.get(f"/admin/intake/formversion/_us117e/schema/{fv.id}/")
        body = r.json()
        sec = body["sections"][0]
        region_q = next(q for q in sec["questions"] if q["name"] == "a0_region")
        subregion_q = next(q for q in sec["questions"] if q["name"] == "a1_subregion")
        # region: a non-empty list with no parent_question.
        assert region_q["parent_question"] == ""
        assert any(o["code"] == "R-TEST" for o in region_q["options"])
        # subregion: declares its parent + each option carries
        # the parent's code so the React cascade can filter.
        assert subregion_q["parent_question"] == "a0_region"
        subregion_opts = [o for o in subregion_q["options"] if o["code"] == "SR-TEST"]
        assert len(subregion_opts) == 1
        assert subregion_opts[0]["parent_code"] == "R-TEST"

    def test_static_preview_links_to_interactive(
        self, _fv_with_roster, django_user_model,
    ):
        """Discovery path — an admin on the static preview should
        see a primary CTA to switch to the interactive one."""
        from django.test import Client
        u = django_user_model.objects.create_user(
            username="ip-staff5", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        r = c.get(
            f"/admin/intake/formversion/_us117b/preview/{_fv_with_roster.id}/",
        )
        body = r.content.decode("utf-8")
        assert f"/admin/intake/formversion/_us117e/preview/{_fv_with_roster.id}/" in body


# --- US-119: rule-pack sync on FormVersion activation ----------------------

class TestRulePackSync:
    """When a FormVersion activates, fan out atomically to DAT-DQA:
    one DqaRule per FormQuestion that has a constraint or skip-logic
    DSL. Idempotent — running sync again upserts (rule_id, version).
    """

    @pytest.fixture
    def _fv_with_constraints(self, db):
        from apps.intake.models import (
            FormConstraint,
            FormQuestion,
            FormSection,
            FormVersion,
        )
        fv = FormVersion.objects.create(
            version=3001, name="sync-test",
            effective_from=date(2026, 1, 1),
            status="pending_approval", author="alice",
        )
        s = FormSection.objects.create(
            form_version=fv, code="C", name="roster", label="Roster", order=1,
        )
        q_age = FormQuestion.objects.create(
            section=s, name="age_years", label="Age", type="integer",
            constraint_expression=". >= 0 and . <= 120",
            constraint_message="age must be 0-120",
            order_in_section=1,
        )
        FormConstraint.objects.create(
            question=q_age,
            dsl={"field": "age_years", "op": "between", "value": [0, 120]},
            message="age must be 0-120",
        )
        q_phone = FormQuestion.objects.create(
            section=s, name="phone", label="Phone", type="text",
            constraint_expression="regex(., '^[+0-9]{9,15}$')",
            constraint_message="phone must be E.164-ish",
            order_in_section=2,
        )
        FormConstraint.objects.create(
            question=q_phone,
            dsl={"field": "phone", "op": "regex", "value": r"^\+[0-9]{9,15}$"},
            message="phone must be E.164-ish",
        )
        # A question with NO constraint — shouldn't produce a rule.
        FormQuestion.objects.create(
            section=s, name="full_name", label="Full name", type="text",
            order_in_section=3,
        )
        return fv

    def test_sync_creates_one_rule_per_constraint(self, _fv_with_constraints):
        from apps.dqa.models import DqaRule
        from apps.intake.rule_pack_sync import sync_rule_pack
        before = DqaRule.objects.count()
        report = sync_rule_pack(_fv_with_constraints, actor="alice")
        after = DqaRule.objects.count()
        assert after == before + 2
        # Returned summary lists what was created.
        assert report["created"] == 2
        assert report["updated"] == 0
        # Naming convention: AC-FORM-<version>-<question_name>
        rule_ids = set(DqaRule.objects.values_list("rule_id", flat=True))
        assert "AC-FORM-3001-age_years" in rule_ids
        assert "AC-FORM-3001-phone" in rule_ids

    def test_sync_links_dsl_into_rule_expression(self, _fv_with_constraints):
        from apps.dqa.models import DqaRule
        from apps.intake.rule_pack_sync import sync_rule_pack
        sync_rule_pack(_fv_with_constraints, actor="alice")
        rule = DqaRule.objects.get(rule_id="AC-FORM-3001-age_years")
        assert rule.expression == {
            "field": "age_years", "op": "between", "value": [0, 120],
        }
        assert rule.error_message_template == "age must be 0-120"
        assert rule.applicability_filter == {
            "entity": "member", "form_version": 3001,
        }

    def test_sync_is_idempotent(self, _fv_with_constraints):
        from apps.dqa.models import DqaRule
        from apps.intake.rule_pack_sync import sync_rule_pack
        sync_rule_pack(_fv_with_constraints, actor="alice")
        count_after_first = DqaRule.objects.count()
        report = sync_rule_pack(_fv_with_constraints, actor="alice")
        # Second pass updates existing rules; no new rows.
        assert DqaRule.objects.count() == count_after_first
        assert report["created"] == 0
        assert report["updated"] >= 2

    def test_sync_skips_questions_without_constraint(self, _fv_with_constraints):
        from apps.dqa.models import DqaRule
        from apps.intake.rule_pack_sync import sync_rule_pack
        sync_rule_pack(_fv_with_constraints, actor="alice")
        # `full_name` has no constraint — no rule emitted.
        assert not DqaRule.objects.filter(
            rule_id="AC-FORM-3001-full_name",
        ).exists()

    def test_sync_emits_audit_event(self, _fv_with_constraints):
        from apps.intake.rule_pack_sync import sync_rule_pack
        from apps.security.models import AuditEvent
        before = AuditEvent.objects.filter(
            entity_type="intake.form_version",
            action="rule_pack_synced",
        ).count()
        sync_rule_pack(_fv_with_constraints, actor="alice")
        after = AuditEvent.objects.filter(
            entity_type="intake.form_version",
            action="rule_pack_synced",
        ).count()
        assert after == before + 1


# --- US-119b: atomic approve + sync ----------------------------------------

class TestApproveFormVersion:
    """approve_form_version() ties the status transition (draft|
    pending_approval → active) and rule-pack sync into one transaction.
    The whole point is to close the gap where v2 looked ACTIVE in
    admin but DAT-DQA never received the rules."""

    @pytest.fixture
    def _pending_fv(self, db):
        from apps.intake.models import (
            FormConstraint,
            FormQuestion,
            FormSection,
            FormVersion,
        )
        fv = FormVersion.objects.create(
            version=4119, name="approve-sync-test",
            effective_from=date(2026, 1, 1),
            status="pending_approval", author="alice",
        )
        s = FormSection.objects.create(
            form_version=fv, code="C", name="roster", label="Roster", order=1,
        )
        q = FormQuestion.objects.create(
            section=s, name="age_years", label="Age", type="integer",
            constraint_expression=". >= 0 and . <= 120",
            constraint_message="age out of range",
            order_in_section=1,
        )
        FormConstraint.objects.create(
            question=q,
            dsl={"field": "age_years", "op": "between", "value": [0, 120]},
            message="age out of range",
        )
        return fv

    def test_approve_transitions_to_active_and_creates_rules(self, _pending_fv):
        from apps.dqa.models import DqaRule
        from apps.intake.services import approve_form_version
        report = approve_form_version(_pending_fv, actor="bob")
        _pending_fv.refresh_from_db()
        assert _pending_fv.status == "active"
        assert _pending_fv.is_active is True
        assert _pending_fv.approved_by == "bob"
        assert _pending_fv.approved_at is not None
        assert report["new_status"] == "active"
        assert report["previous_status"] == "pending_approval"
        # The rule pack landed in DAT-DQA at the same time.
        assert DqaRule.objects.filter(rule_id="AC-FORM-4119-age_years").exists()

    def test_approve_works_from_draft(self, _pending_fv):
        from apps.intake.services import approve_form_version
        _pending_fv.status = "draft"
        _pending_fv.save(update_fields=["status"])
        report = approve_form_version(_pending_fv, actor="bob")
        assert report["previous_status"] == "draft"
        assert report["new_status"] == "active"

    def test_approve_rejected_from_retired(self, _pending_fv):
        from apps.intake.services import FormApprovalError, approve_form_version
        _pending_fv.status = "retired"
        _pending_fv.save(update_fields=["status"])
        with pytest.raises(FormApprovalError, match="cannot be approved"):
            approve_form_version(_pending_fv, actor="bob")

    def test_approve_requires_actor(self, _pending_fv):
        from apps.intake.services import FormApprovalError, approve_form_version
        with pytest.raises(FormApprovalError, match="actor required"):
            approve_form_version(_pending_fv, actor="")

    def test_approve_rollback_on_sync_failure(self, _pending_fv, monkeypatch):
        """If sync_rule_pack raises, the status transition must roll
        back. This is the whole reason the service exists — atomic."""
        from apps.intake import rule_pack_sync as rps
        from apps.intake.models import FormVersion
        from apps.intake.services import approve_form_version

        def boom(*args, **kwargs):
            raise RuntimeError("simulated DAT-DQA failure")

        # The service imports sync_rule_pack from the rule_pack_sync
        # module at call time, so this monkeypatch is picked up.
        monkeypatch.setattr(rps, "sync_rule_pack", boom)
        with pytest.raises(RuntimeError, match="simulated"):
            approve_form_version(_pending_fv, actor="bob")
        # Status must NOT have flipped — atomic rollback.
        fresh = FormVersion.objects.get(pk=_pending_fv.pk)
        assert fresh.status == "pending_approval"
        assert fresh.is_active is False

    def test_approve_emits_audit_event(self, _pending_fv):
        from apps.intake.services import approve_form_version
        from apps.security.models import AuditEvent
        approve_form_version(_pending_fv, actor="bob")
        assert AuditEvent.objects.filter(
            entity_type="intake.form_version",
            action="approve",
            entity_id=_pending_fv.id,
        ).exists()

    def test_admin_action_button_posts_and_redirects(
        self, _pending_fv, django_user_model,
    ):
        from django.test import Client
        u = django_user_model.objects.create_user(
            username="approver", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        r = c.post(
            f"/admin/intake/formversion/_us119b/approve/{_pending_fv.id}/",
        )
        assert r.status_code == 302
        _pending_fv.refresh_from_db()
        assert _pending_fv.status == "active"
        assert _pending_fv.approved_by == "approver"

    def test_admin_action_button_rejects_get(
        self, _pending_fv, django_user_model,
    ):
        """The endpoint is POST-only — GETs from web crawlers /
        accidental link clicks must not change state."""
        from django.test import Client
        u = django_user_model.objects.create_user(
            username="approver2", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        r = c.get(
            f"/admin/intake/formversion/_us119b/approve/{_pending_fv.id}/",
        )
        assert r.status_code == 405


# --- US-S20-001: PII lint --------------------------------------------------

class TestPiiLint:
    """The lint catches NIN-shaped, phone-shaped, email, vehicle-plate
    and 11+-digit values embedded in label / hint / constraint message
    text. Authors often paste a "real-looking" example and we want
    that surfaced before it lands in DqaResult error rows."""

    def test_clean_text_returns_no_violations(self):
        from apps.intake.pii_lint import lint_text
        assert lint_text("Enter a valid phone number.") == []
        assert lint_text("Age must be between 0 and 120.") == []
        assert lint_text("") == []

    def test_detects_uganda_nin_shape(self):
        from apps.intake.pii_lint import lint_text
        v = lint_text("Enter NIN like CM12345678ABCDE")
        assert any(item["rule"] == "nin" and item["matched"] == "CM12345678ABCDE" for item in v)

    def test_detects_uganda_phone(self):
        from apps.intake.pii_lint import lint_text
        v_plus = lint_text("Call +256770123456 for help")
        v_local = lint_text("Try 0770123456 first")
        assert any(item["rule"] == "phone" for item in v_plus)
        assert any(item["rule"] == "phone" for item in v_local)

    def test_detects_email(self):
        from apps.intake.pii_lint import lint_text
        v = lint_text("Contact john.doe@example.com for support")
        assert any(item["rule"] == "email" for item in v)

    def test_detects_long_digit_runs(self):
        """11+ consecutive digits (passport, MM wallet, voter card)."""
        from apps.intake.pii_lint import lint_text
        v = lint_text("Use voter card 12345678901234")
        assert any(item["rule"] == "long_digits" for item in v)

    def test_detects_vehicle_plate(self):
        from apps.intake.pii_lint import lint_text
        v = lint_text("Plate UAE 123J was seen")
        assert any(item["rule"] == "plate" for item in v)

    def test_does_not_double_attribute_phone_as_long_digits(self):
        """A phone-shaped number must classify as 'phone', not also
        get tallied as 'long_digits' — the consumed-span tracking
        in lint_text exists for exactly this case."""
        from apps.intake.pii_lint import lint_text
        v = lint_text("Phone +256770123456 please")
        rules = [item["rule"] for item in v]
        assert "phone" in rules
        assert "long_digits" not in rules

    def test_lint_form_version_walks_the_tree(self, db):
        from apps.intake.models import (
            FormConstraint,
            FormQuestion,
            FormSection,
            FormVersion,
        )
        from apps.intake.pii_lint import lint_form_version
        fv = FormVersion.objects.create(
            version=8001, name="lint-tree",
            effective_from=date(2026, 1, 1),
            status="draft", author="qa",
        )
        s = FormSection.objects.create(
            form_version=fv, code="C", name="ros", label="Ros", order=1,
        )
        q = FormQuestion.objects.create(
            section=s, name="phone", label="Phone",
            type="text", order_in_section=1,
            constraint_message="Enter a phone like +256770123456.",
            hint="Example NIN: CM12345678ABCDE",
        )
        FormConstraint.objects.create(
            question=q, dsl={"field": "phone", "op": "regex", "value": "^\\+?256"},
            message="See example +256770999999 for format.",
        )
        report = lint_form_version(fv)
        assert report["questions_scanned"] == 1
        assert len(report["violations"]) == 1
        entry = report["violations"][0]
        assert entry["section"] == "C" and entry["question"] == "phone"
        rules = {item["rule"] for item in entry["items"]}
        # phone × 2 (hint + constraint message), nin × 1 (hint).
        assert {"phone", "nin"}.issubset(rules)

    def test_admin_pii_lint_button_returns_302(self, db, django_user_model):
        from django.test import Client

        from apps.intake.models import (
            FormQuestion,
            FormSection,
            FormVersion,
        )
        fv = FormVersion.objects.create(
            version=8002, name="lint-admin",
            effective_from=date(2026, 1, 1),
            status="draft", author="qa",
        )
        s = FormSection.objects.create(
            form_version=fv, code="A", name="ident",
            label="Identification", order=1,
        )
        FormQuestion.objects.create(
            section=s, name="nin", label="NIN",
            type="text", order_in_section=1,
            constraint_message="Try CM12345678ABCDE for the format.",
        )
        u = django_user_model.objects.create_user(
            username="linter", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        r = c.post(f"/admin/intake/formversion/_us-pii-lint/{fv.id}/")
        assert r.status_code == 302


# --- US-S20-005: form-versioning hygiene -----------------------------------

class TestFormVersionLockAndClone:
    """Active and retired FormVersions are read-only in the admin —
    operators amend them by cloning to a fresh draft. The clone
    service duplicates the section/question/constraint/skip-logic
    tree at the next version number."""

    @pytest.fixture
    def _populated_active_fv(self, db):
        from apps.intake.models import (
            FormConstraint,
            FormQuestion,
            FormSection,
            FormSkipLogic,
            FormVersion,
        )
        fv = FormVersion.objects.create(
            version=7011, name="lock-clone-fixture",
            effective_from=date(2026, 1, 1),
            status="active", is_active=True, author="qa",
        )
        s = FormSection.objects.create(
            form_version=fv, code="A", name="ident",
            label="Identification", order=1,
        )
        q = FormQuestion.objects.create(
            section=s, name="age", label="Age", type="integer",
            constraint_expression=". >= 0",
            constraint_message="age >= 0",
            order_in_section=1,
        )
        FormConstraint.objects.create(
            question=q, dsl={"field": "age", "op": ">=", "value": 0},
            message="age >= 0",
        )
        FormSkipLogic.objects.create(
            question=q, dsl={"field": "age", "op": "always_true"},
        )
        return fv

    def test_admin_locks_fields_when_active(
        self, _populated_active_fv, django_user_model,
    ):
        from django.contrib.admin.sites import site
        from django.test import RequestFactory

        from apps.intake.admin import FormVersionAdmin
        from apps.intake.models import FormVersion
        u = django_user_model.objects.create_user(
            username="locker", password="p",
            is_staff=True, is_superuser=True,
        )
        rf = RequestFactory().get("/admin/intake/formversion/")
        rf.user = u
        ma = FormVersionAdmin(FormVersion, site)
        ro = ma.get_readonly_fields(rf, _populated_active_fv)
        # `status` must remain editable so a manual retire/reactivate
        # transition is still possible from the admin UI.
        assert "status" not in ro
        # Identity / shape fields are locked.
        for locked in ("name", "version", "description", "effective_from"):
            assert locked in ro, f"expected {locked!r} to be locked when active"

    def test_admin_locks_section_when_parent_locked(
        self, _populated_active_fv, django_user_model,
    ):
        from django.contrib.admin.sites import site
        from django.test import RequestFactory

        from apps.intake.admin import FormSectionAdmin
        from apps.intake.models import FormSection
        u = django_user_model.objects.create_user(
            username="locker2", password="p",
            is_staff=True, is_superuser=True,
        )
        rf = RequestFactory().get("/admin/intake/formsection/")
        rf.user = u
        ma = FormSectionAdmin(FormSection, site)
        sec = _populated_active_fv.sections.first()
        assert ma.has_delete_permission(rf, sec) is False
        ro = ma.get_readonly_fields(rf, sec)
        assert "name" in ro and "label" in ro and "code" in ro

    def test_admin_unlocks_when_draft(self, _populated_active_fv, django_user_model):
        from django.contrib.admin.sites import site
        from django.test import RequestFactory

        from apps.intake.admin import FormVersionAdmin
        from apps.intake.models import FormVersion
        _populated_active_fv.status = "draft"
        _populated_active_fv.save(update_fields=["status"])
        u = django_user_model.objects.create_user(
            username="locker3", password="p",
            is_staff=True, is_superuser=True,
        )
        rf = RequestFactory().get("/admin/intake/formversion/")
        rf.user = u
        ma = FormVersionAdmin(FormVersion, site)
        ro = ma.get_readonly_fields(rf, _populated_active_fv)
        # In draft, fields are editable (only the static class-attr
        # readonly_fields like id/created_at/etc. should appear).
        for editable in ("name", "version", "description", "status"):
            assert editable not in ro

    def test_clone_creates_new_draft_with_next_version(self, _populated_active_fv):
        from apps.intake.models import FormVersion
        from apps.intake.services import clone_form_version
        new_fv = clone_form_version(_populated_active_fv, actor="alice")
        assert new_fv.id != _populated_active_fv.id
        assert new_fv.status == "draft"
        assert new_fv.is_active is False
        assert new_fv.author == "alice"
        assert new_fv.version == _populated_active_fv.version + 1
        # The original is unchanged.
        original = FormVersion.objects.get(pk=_populated_active_fv.id)
        assert original.status == "active"

    def test_clone_duplicates_section_question_constraint_tree(
        self, _populated_active_fv,
    ):
        from apps.intake.services import clone_form_version
        new_fv = clone_form_version(_populated_active_fv, actor="alice")
        assert new_fv.sections.count() == _populated_active_fv.sections.count()
        sec = new_fv.sections.get(name="ident")
        q = sec.questions.get(name="age")
        # Constraints and skip-logic both carried over.
        assert q.constraints.count() == 1
        assert q.skip_logic.count() == 1
        assert q.constraints.first().dsl == {"field": "age", "op": ">=", "value": 0}

    def test_clone_emits_audit_event(self, _populated_active_fv):
        from apps.intake.services import clone_form_version
        from apps.security.models import AuditEvent
        new_fv = clone_form_version(_populated_active_fv, actor="alice")
        assert AuditEvent.objects.filter(
            entity_type="intake.form_version",
            action="clone",
            entity_id=new_fv.id,
        ).exists()

    def test_admin_clone_button_redirects_to_new_draft(
        self, _populated_active_fv, django_user_model,
    ):
        from django.test import Client
        u = django_user_model.objects.create_user(
            username="cloner", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        r = c.post(
            f"/admin/intake/formversion/_us-clone/{_populated_active_fv.id}/",
        )
        assert r.status_code == 302
        # Redirects to the new FormVersion's changeform.
        from apps.intake.models import FormVersion
        new_fv = FormVersion.objects.exclude(pk=_populated_active_fv.id).get()
        assert str(new_fv.id) in r["Location"]


# --- US-S20-004: Kobo push -------------------------------------------------

class TestKoboPush:
    """publish_form_version builds the xlsx, calls KoboConnector
    .publish_xlsform, and persists the returned asset_uid back on
    the FormVersion. We mock the upstream HTTP with `responses` so
    the tests don't need a real Kobo instance."""

    @pytest.fixture
    def _active_fv(self, db):
        from apps.intake.models import (
            FormQuestion,
            FormSection,
            FormVersion,
        )
        fv = FormVersion.objects.create(
            version=6204, name="kobo-push-fixture",
            effective_from=date(2026, 1, 1),
            status="active", is_active=True, author="qa",
        )
        s = FormSection.objects.create(
            form_version=fv, code="A", name="identification",
            label="Identification", order=1,
        )
        FormQuestion.objects.create(
            section=s, name="b1_respondent_name", label="Respondent name",
            type="text", required=True, order_in_section=1,
        )
        return fv

    @pytest.fixture
    def _kobo_setup(self, db):
        """Active KOBO SourceSystem + KoboCredential so the service's
        DB lookup resolves before it ever opens a socket."""
        from apps.ingestion_hub.models import (
            KoboCredential,
            SourceSystem,
            SourceSystemKind,
        )
        src = SourceSystem.objects.create(
            code="KOBO-TEST", kind=SourceSystemKind.KOBO,
            name="Kobo test", is_active=True,
        )
        cred = KoboCredential.objects.create(
            source_system=src,
            server_url="https://kobo.example.test",
            token_encrypted=b"test-token-xyz",
            acquired_by_username="qa",
        )
        return src, cred

    def test_publish_unavailable_when_no_credential(self, _active_fv, db):
        from apps.intake.kobo_push import (
            KoboPushUnavailable,
            publish_form_version,
        )
        with pytest.raises(KoboPushUnavailable):
            publish_form_version(_active_fv, actor="bob")

    def test_publish_refuses_non_active(self, _active_fv, _kobo_setup):
        from apps.intake.kobo_push import KoboPushError, publish_form_version
        _active_fv.status = "draft"
        _active_fv.is_active = False
        _active_fv.save(update_fields=["status", "is_active"])
        with pytest.raises(KoboPushError, match="only active forms"):
            publish_form_version(_active_fv, actor="bob")

    def test_publish_creates_new_asset(self, _active_fv, _kobo_setup):
        """Happy path: no existing kobo_asset_uid, so the connector
        creates a new asset; the returned uid persists on FormVersion."""
        import responses

        from apps.intake.kobo_push import publish_form_version
        server = "https://kobo.example.test"
        with responses.RequestsMock() as rsps:
            rsps.add(
                responses.POST, f"{server}/api/v2/imports/",
                json={"uid": "iABC123"}, status=201,
            )
            rsps.add(
                responses.GET, f"{server}/api/v2/imports/iABC123/",
                json={
                    "status": "complete",
                    "messages": {"created": [{"uid": "aDEPLOY9"}]},
                },
                status=200,
            )
            rsps.add(
                responses.POST,
                f"{server}/api/v2/assets/aDEPLOY9/deployment/",
                json={"active": True}, status=200,
            )
            report = publish_form_version(_active_fv, actor="bob")
        assert report["status"] == "complete"
        assert report["asset_uid"] == "aDEPLOY9"
        assert report["deployed"] is True
        _active_fv.refresh_from_db()
        assert _active_fv.kobo_asset_uid == "aDEPLOY9"

    def test_publish_replaces_existing_asset(self, _active_fv, _kobo_setup):
        """When FormVersion has a kobo_asset_uid, the connector posts
        with `destination` pointing at the existing asset so Kobo
        treats it as a new version of the same form."""
        import json

        import responses

        from apps.intake.kobo_push import publish_form_version
        _active_fv.kobo_asset_uid = "aOLD777"
        _active_fv.save(update_fields=["kobo_asset_uid"])
        server = "https://kobo.example.test"
        with responses.RequestsMock() as rsps:
            rsps.add(
                responses.POST, f"{server}/api/v2/imports/",
                json={"uid": "iREPLACE1"}, status=201,
            )
            rsps.add(
                responses.GET, f"{server}/api/v2/imports/iREPLACE1/",
                json={
                    "status": "complete",
                    "messages": {"updated": [{"uid": "aOLD777"}]},
                },
                status=200,
            )
            rsps.add(
                responses.POST,
                f"{server}/api/v2/assets/aOLD777/deployment/",
                json={"active": True}, status=200,
            )
            report = publish_form_version(_active_fv, actor="bob")
            # Inspect the upload call to confirm `destination` carried
            # the existing asset URL — this is the whole point of
            # the replace path (preserves submission history).
            upload_call = next(c for c in rsps.calls if c.request.url.endswith("/imports/"))
            body = upload_call.request.body
            # multipart body is bytes; the `destination` field is
            # present somewhere in the payload.
            payload = body.decode("utf-8", errors="ignore") if isinstance(body, bytes) else json.dumps(body)
            assert "aOLD777" in payload
        assert report["asset_uid"] == "aOLD777"

    def test_publish_audits_even_on_timeout(self, _active_fv, _kobo_setup):
        """The import task can stall — we still audit the attempt so
        an operator can reconcile with Kobo's UI."""
        import responses

        from apps.intake.kobo_push import publish_form_version
        from apps.security.models import AuditEvent
        server = "https://kobo.example.test"
        with responses.RequestsMock() as rsps:
            rsps.add(
                responses.POST, f"{server}/api/v2/imports/",
                json={"uid": "iSTALL5"}, status=201,
            )
            # Always-processing — exhaust the poll budget.
            rsps.add(
                responses.GET, f"{server}/api/v2/imports/iSTALL5/",
                json={"status": "processing"}, status=200,
            )
            report = publish_form_version(
                _active_fv, actor="bob",
                connector=_FastPollKoboConnector(),
            )
        assert report["status"] == "timeout"
        assert AuditEvent.objects.filter(
            entity_type="intake.form_version",
            action="kobo_publish",
            entity_id=_active_fv.id,
        ).exists()

    def test_admin_button_posts(self, _active_fv, _kobo_setup, django_user_model):
        import responses
        from django.test import Client

        u = django_user_model.objects.create_user(
            username="publisher", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        server = "https://kobo.example.test"
        with responses.RequestsMock() as rsps:
            rsps.add(responses.POST, f"{server}/api/v2/imports/",
                     json={"uid": "iADMIN"}, status=201)
            rsps.add(responses.GET, f"{server}/api/v2/imports/iADMIN/",
                     json={"status": "complete",
                           "messages": {"created": [{"uid": "aADMIN9"}]}},
                     status=200)
            rsps.add(responses.POST,
                     f"{server}/api/v2/assets/aADMIN9/deployment/",
                     status=200, json={"active": True})
            r = c.post(f"/admin/intake/formversion/_uskobo/push/{_active_fv.id}/")
        assert r.status_code == 302
        _active_fv.refresh_from_db()
        assert _active_fv.kobo_asset_uid == "aADMIN9"


class _FastPollKoboConnector:
    """A test connector that polls 3× at zero interval so the
    timeout-path test finishes in milliseconds."""

    def publish_xlsform(self, credentials, **kwargs):
        from apps.ingestion_hub.connectors.kobo import KoboConnector
        return KoboConnector().publish_xlsform(
            credentials, **{**kwargs, "poll_attempts": 3, "poll_interval_s": 0},
        )


# --- US-117d: in-admin HTML preview ----------------------------------------

class TestFormVersionPreview:
    """Preview view renders the FormVersion as HTML — sections as
    cards, questions as disabled controls + annotation strip."""

    def _staff_client(self, db, django_user_model):
        from django.test import Client
        u = django_user_model.objects.create_user(
            username="prev-staff", password="p",
            is_staff=True, is_superuser=True,
        )
        c = Client()
        c.force_login(u)
        return c

    @pytest.fixture
    def _seeded_form(self, db):
        from apps.intake.models import (
            FormQuestion,
            FormSection,
            FormVersion,
        )
        from apps.reference_data.models import ChoiceList, ChoiceOption
        fv = FormVersion.objects.create(
            version=4000, name="preview-test",
            effective_from=date(2026, 1, 1),
            status="draft", author="alice",
        )
        # ChoiceList for a select_one question.
        cl = ChoiceList.objects.create(
            list_name="preview_sex", version=1,
            effective_from=date(2026, 1, 1),
            status="active", author="alice", approved_by="bob",
        )
        ChoiceOption.objects.create(
            choice_list=cl, code="M", label="Male",
            language="en", sort_order=1,
        )
        ChoiceOption.objects.create(
            choice_list=cl, code="F", label="Female",
            language="en", sort_order=2,
        )
        # Section + question variety.
        a = FormSection.objects.create(
            form_version=fv, code="A", name="ident", label="Identification", order=1,
        )
        FormQuestion.objects.create(
            section=a, name="full_name", label="Full name", type="text",
            required=True, hint="Surname, then first name", order_in_section=1,
        )
        FormQuestion.objects.create(
            section=a, name="sex", label="Sex", type="select_one",
            choice_list_ref=cl, required=True, order_in_section=2,
        )
        FormQuestion.objects.create(
            section=a, name="age_years", label="Age in years", type="integer",
            constraint_expression=". >= 0 and . <= 120",
            constraint_message="age must be 0-120",
            order_in_section=3,
        )
        FormQuestion.objects.create(
            section=a, name="age_hidden", label="Hidden if no age",
            type="text", relevant_expression="${age_years} != ''",
            order_in_section=4,
        )
        FormQuestion.objects.create(
            section=a, name="gps", label="GPS", type="geopoint",
            order_in_section=5,
        )
        return fv

    def test_preview_returns_200(self, _seeded_form, db, django_user_model):
        c = self._staff_client(db, django_user_model)
        r = c.get(f"/admin/intake/formversion/_us117b/preview/{_seeded_form.id}/")
        assert r.status_code == 200

    def test_preview_renders_section_and_questions(
        self, _seeded_form, db, django_user_model,
    ):
        c = self._staff_client(db, django_user_model)
        r = c.get(f"/admin/intake/formversion/_us117b/preview/{_seeded_form.id}/")
        body = r.content.decode()
        # Section heading + question labels.
        assert "Identification" in body
        assert "Full name" in body
        assert "Age in years" in body
        # Required marker.
        assert "*" in body
        # Hint text rendered.
        assert "Surname, then first name" in body
        # select_one renders the choice list options.
        assert "Male" in body and "Female" in body
        # geopoint placeholder shown.
        assert "latitude" in body or "GPS" in body
        # Annotation strip surfaces relevant + constraint.
        assert "relevant:" in body
        assert "constraint:" in body

    def test_preview_has_download_link(
        self, _seeded_form, db, django_user_model,
    ):
        c = self._staff_client(db, django_user_model)
        r = c.get(f"/admin/intake/formversion/_us117b/preview/{_seeded_form.id}/")
        body = r.content.decode()
        # Header has the XLSForm download link + the interactive
        # preview CTA (US-117e replaced the external Kobo link).
        assert "Download XLSForm" in body
        assert f"export-xlsform/{_seeded_form.id}/" in body
        assert f"_us117e/preview/{_seeded_form.id}/" in body

    def test_change_form_links_to_preview_and_download(
        self, _seeded_form, db, django_user_model, settings,
    ):
        settings.QUESTIONNAIRE_EDITOR_V2 = True
        c = self._staff_client(db, django_user_model)
        r = c.get(f"/admin/intake/formversion/{_seeded_form.id}/change/")
        body = r.content.decode()
        # Action bar wires all three preview/download paths.
        assert "Interactive preview" in body
        assert "Static preview" in body
        assert "Download XLSForm" in body
        assert f"_us117e/preview/{_seeded_form.id}/" in body

    def test_preview_accepts_version_number_as_fallback(
        self, _seeded_form, db, django_user_model,
    ):
        """`/preview/<version>/` works as a fallback for operators
        who type the URL by hand — FormVersion has ULID primary
        keys, but `1` is what they see on the changelist."""
        c = self._staff_client(db, django_user_model)
        # _seeded_form has version=4000; lookup by version number
        # resolves to the same row.
        r = c.get(
            f"/admin/intake/formversion/_us117b/preview/{_seeded_form.version}/",
        )
        assert r.status_code == 200
        body = r.content.decode()
        assert "Identification" in body

    def test_preview_404_on_unknown_id(self, db, django_user_model):
        c = self._staff_client(db, django_user_model)
        r = c.get("/admin/intake/formversion/_us117b/preview/999999/")
        assert r.status_code == 404

    def test_export_accepts_version_number_as_fallback(
        self, _seeded_form, db, django_user_model,
    ):
        """Same fallback for the US-118 download URL."""
        c = self._staff_client(db, django_user_model)
        r = c.get(
            f"/admin/intake/formversion/_us118/export-xlsform/{_seeded_form.version}/",
        )
        assert r.status_code == 200
        assert r.content[:2] == b"PK"
